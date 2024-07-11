import sys
import csv
from PyQt6.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget, QMenuBar, QMenu, QFileDialog, QMessageBox
from PyQt6.QtGui import QAction
import openpyxl

class SpreadsheetApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("PyQt6 Spreadsheet")
        self.setGeometry(100, 100, 800, 600)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)

        self.create_menu_bar()
        self.create_table_widget()

    def create_menu_bar(self):
        menu_bar = QMenuBar(self)
        self.setMenuBar(menu_bar)

        file_menu = QMenu("&File", self)
        menu_bar.addMenu(file_menu)

        new_action = QAction("&New", self)
        new_action.triggered.connect(self.new_file)
        file_menu.addAction(new_action)

        open_action = QAction("&Open", self)
        open_action.triggered.connect(self.open_file)
        file_menu.addAction(open_action)

        save_action = QAction("&Save", self)
        save_action.triggered.connect(self.save_file)
        file_menu.addAction(save_action)

    def create_table_widget(self):
        self.table = QTableWidget(50, 26)  # 50 rows, 26 columns (A-Z)
        self.layout.addWidget(self.table)

        # Set column headers (A, B, C, ...)
        headers = [chr(65 + i) for i in range(26)]
        self.table.setHorizontalHeaderLabels(headers)

        # Connect cell changed signal
        self.table.cellChanged.connect(self.cell_changed)

    def cell_changed(self, row, column):
        # Here you can implement cell value calculation, formatting, etc.
        pass

    def new_file(self):
        self.table.clearContents()
        self.table.setRowCount(50)
        self.table.setColumnCount(26)

    def open_file(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Spreadsheet", "", "Spreadsheet Files (*.csv *.xlsx);;All Files (*)")
        if file_name:
            try:
                if file_name.endswith('.csv'):
                    self.open_csv(file_name)
                elif file_name.endswith('.xlsx'):
                    self.open_xlsx(file_name)
                else:
                    raise ValueError("Unsupported file format")
                
                QMessageBox.information(self, "File Opened", f"Successfully opened {file_name}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"An error occurred while opening the file: {str(e)}")

    def open_csv(self, file_name):
        with open(file_name, 'r', newline='') as csvfile:
            csv_reader = csv.reader(csvfile)
            data = list(csv_reader)
            
            row_count = len(data)
            col_count = len(max(data, key=len))
            
            self.table.setRowCount(max(row_count, 50))
            self.table.setColumnCount(max(col_count, 26))
            
            for row, row_data in enumerate(data):
                for col, cell_data in enumerate(row_data):
                    item = QTableWidgetItem(cell_data)
                    self.table.setItem(row, col, item)

    def open_xlsx(self, file_name):
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        
        row_count = sheet.max_row
        col_count = sheet.max_column
        
        self.table.setRowCount(max(row_count, 50))
        self.table.setColumnCount(max(col_count, 26))
        
        for row in range(1, row_count + 1):
            for col in range(1, col_count + 1):
                cell_data = sheet.cell(row=row, column=col).value
                if cell_data is not None:
                    item = QTableWidgetItem(str(cell_data))
                    self.table.setItem(row - 1, col - 1, item)

    def save_file(self):
        file_name, _ = QFileDialog.getSaveFileName(self, "Save Spreadsheet", "", "CSV Files (*.csv);;Excel Files (*.xlsx);;All Files (*)")
        if file_name:
            try:
                if file_name.endswith('.csv'):
                    self.save_csv(file_name)
                elif file_name.endswith('.xlsx'):
                    self.save_xlsx(file_name)
                else:
                    raise ValueError("Unsupported file format")
                
                QMessageBox.information(self, "File Saved", f"Successfully saved to {file_name}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"An error occurred while saving the file: {str(e)}")

    def save_csv(self, file_name):
        with open(file_name, 'w', newline='') as csvfile:
            csv_writer = csv.writer(csvfile)
            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item is not None:
                        row_data.append(item.text())
                    else:
                        row_data.append('')
                csv_writer.writerow(row_data)

    def save_xlsx(self, file_name):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item is not None:
                    sheet.cell(row=row+1, column=col+1, value=item.text())
        
        workbook.save(file_name)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = SpreadsheetApp()
    window.show()
    sys.exit(app.exec())